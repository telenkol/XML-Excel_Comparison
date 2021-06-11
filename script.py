# McMaster University - Office of the Registrar
# Date: May 27, 2021
# Python3.9

# Test Python script for XML file parsing. It can currently parse OUAC XML files for program names and student numbers,
# then export that data to a CSV

# Imports
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
# Import for progressbar
from tqdm import tqdm
import time
# Colours for console
from colorama import Fore
# Import the os module
import os

# Global variable declaration
xml_applicants = {}
excel_applicants = {}
xml_file = ""
excel_file = ""

# Function to parse XML file
def parse_xml(filepath):
    # Load global variables
    global xml_applicants

    # Load file from disk
    tree = ET.parse(filepath)

    # Find root node
    root = tree.getroot()

    # Iterate over XML file and pull data about applicants
    for student in tqdm(root.findall('.//Applicant')):
        # Get student number
        student_number = student.find(
            './Application/ApplicationDecisionDetail/UniversityAssignedIdentifier').text

        # Get program name
        program = student.findall(
            './Application/ApplicationDecisionDetail/ApplicationDegreeProgram/ApplicationQuestion/AnswerText')[1].text

        # Check if there is an entry for the applicant in the dictionary. If there is none, create a new entry,
        # otherwise update the existing one.
        if student_number not in xml_applicants:
            # Create a new array with their program
            xml_applicants[student_number] = [program]
        else:
            # Add the current program to the student's array
            xml_applicants[student_number].append(program)

# Function to parse Excel file
def parse_excel(filepath):
    # Load global variables
    global excel_applicants

    # Load workbook and get correct sheet (first)
    book = load_workbook(filename=filepath)
    sheet = book[book.sheetnames[0]]

    # Row counter, starts at row 3
    index = 3

    while sheet.cell(row=index, column=6).value is not None:

        # Get student number
        student_number = sheet.cell(row=index, column=8).value

        # Get program
        program = sheet.cell(row=index, column=4).value

        # Check if there is an entry for the applicant in the dictionary. If there is none, create a new entry,
        # otherwise update the existing one.
        if student_number not in excel_applicants:
            # Create a new array with their program
            excel_applicants[student_number] = [program]
        else:
            # Add the current program to the student's array
            excel_applicants[student_number].append(program)

        index += 1  # Iterate row count


# Exports the given data to file
def export_csv(filepath, data):
    # Create/open file with write permissions
    try:
        csv_export = open(filepath, "w")
    except (PermissionError):
        printError("Could not write to output file. Please ensure that the file is not currently and try again.")

    # Write to disk
    csv_export.write(data)

    # Close file
    csv_export.close()


# Taken from https://thispointer.com/python-check-if-two-lists-are-equal-or-not-covers-both-ordered-unordered-lists/
def identical_lists(list_1, list_2):
    """ Check if both the lists are of same length and if yes then compare
    sorted versions of both the list to check if both of them are equal
    i.e. contain similar elements with same frequency. """
    if len(list_1) != len(list_2):
        return False
    return sorted(list_1) == sorted(list_2)


# Checks if the student has the same offers in both files
# Returns the csv data of the students with discrepancies
def compare_applicants_data():
    data = ""
    for (student, xml_offers) in tqdm(xml_applicants.items()):
        # Offers in Excel file
        excel_offers = excel_applicants.get(student)

        # Different offers in one of the files
        if xml_offers is not None and excel_offers is not None and not identical_lists(xml_offers, excel_offers):
            # Fill student number cell
            temp = str(student)
            # Discrepancy type
            temp += ", " + "Different offers"
            # Go to next row
            data += temp + "\n"

        # Student number is not in one of the files
        elif xml_offers is None:
            # Fill student number cell
            temp = str(student)
            # Discrepancy type
            temp += ", " + "Student number not in XML File"
            # Go to next row
            data += temp + "\n"

        # Student number is not in one of the files
        elif excel_offers is None:
            # Fill student number cell
            temp = str(student)
            # Discrepancy type
            temp += ", " + "Student number not in Excel File"
            # Go to next row
            data += temp + "\n"

    return data

# Helper methods for printing to console
def printTitle(title):
    print(Fore.YELLOW + "\n[" + title + "]" + Fore.WHITE)

def printError(error):
    print(Fore.RED + "[ERROR] " + Fore.WHITE + error)
    input("\nPress the enter key to exit.")
    quit()

# Main method
def main():

    # Get the current working directory
    global xml_file, excel_file
    cwd = os.getcwd()

    printTitle("Looking for .xml and .xlsx files")
   
    # Find xml and excel file names
    for filename in os.listdir(os.path.dirname(os.path.abspath(__file__))):
        if filename.endswith('.xml') and not filename.startswith('~$'):
            xml_file = os.path.dirname(os.path.abspath(__file__)) + "\\" + filename
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            excel_file = os.path.dirname(os.path.abspath(__file__)) + "\\" + filename

    # Verify files were found
    if (excel_file == ""):
        printError("No .xlsx file was found.")
    elif(xml_file == ""):
        printError("No .xml file was found.")
    else:
        print("Both files found!")
        
    printTitle("Parsing data from XML file")
    parse_xml(xml_file)

    printTitle("Parsing data from Excel file")
    parse_excel(excel_file)

    printTitle("Comparing files")
    export_csv(os.path.dirname(os.path.abspath(__file__)) + "\\discrepancies.csv", compare_applicants_data())
    # compare_applicants_data()

    # Print confirmation message and wait to exit program
    print("\nProgram ran successfully, the results are saved as " + Fore.YELLOW + "'discrepancies.csv'\n" + Fore.WHITE)
    input("Press the enter key to exit.")

main()